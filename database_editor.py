import mysql.connector
from tkinter import *
from tkinter import messagebox
import basic_operations
import sqlite3
from tkinter import ttk
import os
from os.path import join
from openpyxl import Workbook, load_workbook
import datetime
from openpyxl.styles import Font


def typeea(value):
	if value == "$ per item":
		return 0
	if value == "$ per weight":
		return 1
	if value == 0:
		return "each"
	if value == 1:
		return "Kg"
def write(path):
    print('Creating a new file')
    print(path)
    name = 'file'  # Name of text file coerced with +.txt

    file = open(join(path, name),'w')   # Trying to create a new file or open one
    file.close()



class create_item():
	def __init__(self):
		self.vonn = sqlite3.connect("LocalData")
		self.v = self.vonn.cursor()
		self.v.execute("SELECT * FROM host_connect_info")
		self.host = self.v.fetchall()[0]
		self.root = Tk()
		self.root.title("Create Item")
		self.root.geometry("500x400+0+0")
	
		self.title_label = Label(self.root, text = "Create Item", font = ("Arial", 35))
		self.title_label.place(anchor=CENTER, relx=0.5, rely=0.15)

		self.plu_label = Label(self.root, text = "PLU*:")
		self.plu_label.place(anchor=E, relx=0.38, rely = 0.3)

		self.description_label = Label(self.root, text = "Description*:")
		self.description_label.place(anchor=E, relx=0.38, rely = 0.37)

		self.price_label = Label(self.root, text = "Price: $")
		self.price_label.place(anchor=E, relx=0.38, rely = 0.44)

		self.type_label = Label(self.root, text = "Type:")
		self.type_label.place(anchor=E, relx=0.38, rely = 0.51)

		self.category_label = Label(self.root, text = "Category:") 
		self.category_label.place(anchor=E, relx=0.38, rely = 0.58)

		self.points_label = Label(self.root, text = "Bonus Points:")
		self.points_label.place(anchor=E, relx=0.38, rely = 0.65)


		#entrys

		self.plu_entry = Entry(self.root)
		self.plu_entry.place(anchor=W, relx=0.4, rely = 0.3)
		self.plu_entry.focus()


		self.description_entry = Entry(self.root)
		self.description_entry.place(anchor=W, relx=0.4, rely = 0.37)

		self.price_entry = Entry(self.root)
		self.price_entry.place(anchor=W, relx=0.4, rely = 0.44)

		self.typea = StringVar()
		self.typea.set("$ per item")

		self.options = ["$ per item", "$ per weight"]
		
		self.type_entry = Entry(self.root)
		self.type_entry.place(anchor=W, relx = 0.4, rely=0.51)
	
		

		
		self.category_entry = Entry(self.root)
		self.category_entry.place(anchor=W, relx=0.4, rely = 0.58)

		

		self.points_entry = Entry(self.root)
		self.points_entry.place(anchor=W, relx=0.4, rely = 0.65)

		self.submit_button = Button(self.root, text="Create", command=self.create)
		self.submit_button.place(anchor=CENTER, relx=0.5, rely=0.8)

		self.close_button = Button(self.root, text="❌", command=self.close)
		self.close_button.place(anchor=CENTER, relx=0.5, rely=0.9)

		self.root.bind("<KeyRelease>", lambda a:self.keyrelease())

		self.root.bind("<Return>", lambda a:self.create())
		self.root.mainloop()
		return None
	def keyrelease(self):
		
		if self.type_entry.get() != "0":
			
			if self.type_entry.get() != "1":
				self.type_entry.delete(0, END)
	def close(self):
		self.root.destroy()
		self.root.quit()

	def create(self):
		if self.plu_entry.get() == "":
			self.plu_entry.focus()
		else:
			if self.description_entry.get() == "":
				self.description_entry.focus()
			else:
				if self.price_entry.get() == "":
					self.price_entry.focus()
				else:
					if self.type_entry.get() == "":
						self.type_entry.focus()
					else:
						if self.category_entry.get() == "":
							self.category_entry.focus()
						else:
							if self.points_entry.get() == "":
								self.points_entry.focus()
							else:
								self.plu = self.plu_entry.get()
								self.description = self.description_entry.get()
								self.price = self.price_entry.get()
								self.category = self.category_entry.get()
								self.points = self.points_entry.get()
								if basic_operations.is_float(self.price) == True and basic_operations.is_int(self.points) == True:

									
									self.conn = mysql.connector.connect(
										host=self.host[0],
										user=self.host[1],
										passwd=self.host[2],
										database=self.host[3]
										
									)
									self.c = self.conn.cursor()
									self.c.execute("SELECT * FROM items WHERE plu = %s", [self.plu])
									self.data = self.c.fetchall()
									
									if len(self.data)>0:
										self.plu_entry.delete(0, END)
										self.plu_entry.focus()
										Label(self.root, text="PLU in use", fg = 'red').place(anchor=CENTER, relx=0.5, rely=0.72)
									else:
										self.c.execute("INSERT INTO items(plu, description, price, type, category, points, quantity) VALUES(%s,%s,%s,%s,%s,%s, 0)", [self.plu, self.description, self.price, self.type_entry.get(), self.category, self.points])
										self.conn.commit()
										self.root.destroy()
										self.root.quit()
								elif basic_operations.is_float(self.price) == True:
									self.points_entry.focus()
								else:
									self.price_entry.focus()

		return True
class edit_item():
	def __init__(self, plu):
		self.plu_direct = plu
		

		self.vonn = sqlite3.connect("LocalData")
		self.v = self.vonn.cursor()
		self.v.execute("SELECT * FROM host_connect_info")
		self.host = self.v.fetchall()[0]
		self.conn = mysql.connector.connect(
									host=self.host[0],
									user=self.host[1],
									passwd=self.host[2],
									database=self.host[3]
									
								)
		self.c = self.conn.cursor()
		self.root = Tk()
		self.root.title("Edit Item")
		self.root.geometry("500x400+0+0")
	
		self.title_label = Label(self.root, text = "Edit Item", font = ("Arial", 35))
		self.title_label.place(anchor=CENTER, relx=0.5, rely=0.15)

		self.plu_label = Label(self.root, text = "PLU*:")
		self.plu_label.place(anchor=E, relx=0.38, rely = 0.3)

		self.description_label = Label(self.root, text = "Description*:")
		self.description_label.place(anchor=E, relx=0.38, rely = 0.37)

		self.price_label = Label(self.root, text = "Price: $")
		self.price_label.place(anchor=E, relx=0.38, rely = 0.44)

		self.type_label = Label(self.root, text = "Type:")
		self.type_label.place(anchor=E, relx=0.38, rely = 0.51)

		self.category_label = Label(self.root, text = "Category:") 
		self.category_label.place(anchor=E, relx=0.38, rely = 0.58)

		self.points_label = Label(self.root, text = "Bonus Points:")
		self.points_label.place(anchor=E, relx=0.38, rely = 0.65)

		self.quantity_label = Label(self.root, text = "Quantity In Stock:")
		self.quantity_label.place(anchor=E, relx=0.38, rely = 0.72)


		#entrys
		#
		self.search_button = Button(self.root, text = "🔍")
		self.search_button.place(anchor=CENTER, relx=0.85, rely = 0.293)

		self.plu_entry = Entry(self.root)
		self.plu_entry.place(anchor=W, relx=0.4, rely = 0.3)
		self.plu_entry.focus()

		self.description_entry = Entry(self.root)
		self.description_entry.place(anchor=W, relx=0.4, rely = 0.37)

		self.price_entry = Entry(self.root)
		self.price_entry.place(anchor=W, relx=0.4, rely = 0.44)



		self.type_entry = Entry(self.root)
		self.type_entry.place(anchor=W, relx = 0.4, rely=0.51)

		
		self.category_entry = Entry(self.root)
		self.category_entry.place(anchor=W, relx=0.4, rely = 0.58)

		self.points_entry = Entry(self.root)
		self.points_entry.place(anchor=W, relx=0.4, rely = 0.65)

		self.quantity_entry = Entry(self.root)
		self.quantity_entry.place(anchor=W, relx=0.4, rely = 0.72)

		self.save_button = Button(self.root, text="Save", command=self.save)
		self.save_button.place(anchor=CENTER, relx=0.5, rely=0.8)

		self.close_button = Button(self.root, text="❌", command=self.close)
		self.close_button.place(anchor=CENTER, relx=0.5, rely=0.9)

		if self.plu_direct != None:
			self.plu_entry.insert(0, self.plu_direct)
			self.search()

		self.root.bind("<KeyRelease>", lambda a:self.keyrelease())
		self.root.bind("<Return>", lambda a:self.enter())
		self.root.mainloop()
		return None
	def keyrelease(self):
		
	
		if self.type_entry.get() != "0":
			
			if self.type_entry.get() != "1":
				self.type_entry.delete(0, END)	
	def close(self):
		self.root.destroy()
		self.root.quit()

	def enter(self):
		if self.plu_entry.get() != "" and self.description_entry.get() == "":
			self.search()
		else:
			if self.plu_entry.get() == "":
				self.plu_entry.focus()
			else:
				if self.description_entry.get() == "":
					self.description_entry.focus()
				else:
					if self.price_entry.get() == "":
						self.price_entry.focus()
					else:
						if self.category_entry.get() == "":
							self.category_entry.focus()
						else:
							if self.points_entry.get() == "":
								self.points_entry.focus()
							else:
								self.save()

	def search(self):
		self.c.execute("SELECT * FROM items WHERE plu = %s", [self.plu_entry.get()])
		self.data = self.c.fetchall()
		if self.data == []:
			self.plu_entry.delete(0, END)
			
		else:
			print(self.data)
			self.plu = self.plu_entry.get()
			self.refil_entry(self.description_entry, self.data[0][1])
			self.refil_entry(self.price_entry, self.data[0][5])
			self.refil_entry(self.category_entry, self.data[0][3])
			self.refil_entry(self.points_entry, self.data[0][4])
			self.refil_entry(self.type_entry, self.data[0][2])
			self.refil_entry(self.quantity_entry, self.data[0][6])
		return True
	def save(self):
		if self.plu_entry.get() != "":
			self.c.execute("SELECT * FROM items WHERE plu = %s", [self.plu_entry.get()])
			self.data = self.c.fetchall()
			
				
			if basic_operations.is_int(self.points_entry.get()) == True:
				if basic_operations.is_float(self.price_entry.get()) == True:
					if basic_operations.is_float(self.quantity_entry.get())==True:
						
						self.c.execute("UPDATE items SET plu = %s, description = %s, price = %s, type = %s, category = %s, points = %s, quantity = %s WHERE plu = %s",
								[
								self.plu_entry.get(),
								self.description_entry.get(), 
								self.price_entry.get(), 
								self.type_entry.get(), 
								self.category_entry.get(), 
								self.points_entry.get(), 
								self.quantity_entry.get(), 
								self.plu
								]
								)
						self.conn.commit()
						self.root.destroy()
						self.root.quit()
					else:
						self.quantity_entry.delete(0, END)
						self.quantity_entry.focus()
				else:
					self.price_entry.delete(0, END)
					self.price_entry.focus()
			else:
				self.points_entry.delete(0, END)
				self.points_entry.focus()
		return True
	def refil_entry(self, entry, inputa):
		entry.delete(0, END)
		entry.insert(0, inputa)
class delete_item():
	def __init__(self,host):
		self.vonn = sqlite3.connect("LocalData")
		self.v = self.vonn.cursor()
		self.v.execute("SELECT * FROM host_connect_info")
		self.host = self.v.fetchall()[0]
		self.conn = mysql.connector.connect(
									host=self.host[0],
									user=self.host[1],
									passwd=self.host[2],
									database=self.host[3]
									
								)
		self.c = self.conn.cursor()
		self.root = Tk()
		self.root.title("Delete Item")
		self.root.geometry("500x400+0+0")
	
		self.title_label = Label(self.root, text = "Delete Item", font = ("Arial", 35))
		self.title_label.place(anchor=CENTER, relx=0.5, rely=0.15)

		self.plu_label = Label(self.root, text = "Search PLU", font = ("Arial", 15))
		self.plu_label.place(anchor=CENTER, relx=0.5, rely=0.3)

		self.search_button = Button(self.root, text = "🔍", command = self.search)
		self.search_button.place(anchor=CENTER, relx=0.75, rely = 0.1+0.293)
		self.delete_button = Button(self.root, text = "🗑️", command = self.delete)
		self.delete_button.place(anchor=CENTER, relx=0.5, rely = 0.8)

		self.info_label = Label(self.root, text= "PLU: \nDescription: \nPrice: \nCategory: \nQuantity: \nBonus Points: ", justify=RIGHT)
		self.info_label.place(anchor = E, relx=0.5, rely=0.6)

		self.result_label = Label(self.root, text= "", justify=LEFT)
		self.result_label.place(anchor = W, relx=0.5, rely=0.6)

		self.close_button = Button(self.root, text="❌", command=self.close)
		self.close_button.place(anchor=CENTER, relx=0.5, rely=0.9)

		self.plu_entry = Entry(self.root)
		self.plu_entry.place(anchor=CENTER, relx=0.5, rely = 0.4)
		self.plu_entry.focus()

		self.root.bind("<Return>", lambda a: self.search())
		self.root.mainloop()

	def close(self):
		self.root.destroy()
		self.root.quit()

	def search(self):
		self.c.execute("SELECT * FROM items WHERE plu = %s", [self.plu_entry.get()])
		self.data = self.c.fetchall()
		if self.data == []:
			self.plu_entry.delete(0, END)
			self.result_label.configure(text="")
			return False

		else:
			self.result_label.configure(text = f"{self.data[0][0]}\n{self.data[0][1]}\n{self.data[0][5]}\n{self.data[0][3]}\n{self.data[0][6]}\n{self.data[0][4]}")
			return True
	def delete(self):
		if self.search() == True:
			if messagebox.askquestion ("Delete Item", "Are you sure you want to delete this item?", icon = 'warning') == "yes":
				self.c.execute("DELETE FROM items WHERE plu = %s", [self.plu_entry.get()])
				self.conn.commit()
				self.close()
			else:
				self.plu_entry.delete(0, END)
				self.result_label.configure(text="")
		return True

class create_user():
	def __init__(self):
		self.vonn = sqlite3.connect("LocalData")
		self.v = self.vonn.cursor()
		self.v.execute("SELECT * FROM host_connect_info")
		self.host = self.v.fetchall()[0]
		
		self.root = Tk()
		self.root.title("Create User")
		self.root.geometry("750x600+0+0")
	
		self.title_label = Label(self.root, text = "Create User", font = ("Arial", 35))
		self.title_label.place(anchor=CENTER, relx=0.35, rely=0.15)


		Label(self.root, text = "ID:").place(anchor=E, relx=0.2, rely = 0.3)

		Label(self.root, text = "Password:").place(anchor=E, relx=0.2, rely = 0.37)

		Label(self.root, text = "Name:").place(anchor=E, relx=0.2, rely = 0.44)


		Label(self.root, text = "Permission Presets:").place(anchor=E, relx=0.2, rely = 0.49)





		#entrys

		self.id_entry = Entry(self.root, width = 27)
		self.id_entry.place(anchor=W, relx=0.22, rely = 0.3)
		self.id_entry.focus()

		self.password_entry = Entry(self.root, width = 27)
		self.password_entry.place(anchor=W, relx=0.22, rely = 0.37)

		self.name_entry = Entry(self.root, width = 27)
		self.name_entry.place(anchor=W, relx=0.22, rely = 0.44)
		self.perms = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
		

		self.options = [
		    "Delete Local Data",
		    "Edit Items",
		    "Edit Users",
		    "Edit Accounts",
		    "Transactions",
		    "Local Settings",
		    "Global Settings",
		    "Statisitics"]	
		Label(self.root,text="Permissions").place(anchor=W, relx = 0.65, rely = 0.08)
		self.permission_list = Listbox(self.root, height=30, width = 25)
		self.permission_list.place(anchor = NW, relx = 0.65, rely = 0.1)

		self.toggle_permissions_button = Button(self.root, text="Toggle Selected Permission", command = self.toggle_permissions)
		self.toggle_permissions_button.place(anchor=CENTER, relx= 0.35, rely = 0.6)

		self.preset_admin_button = Button(self.root, text = "Admin", command = self.presetadmin)
		self.preset_admin_button.place(anchor =W, relx = 0.22, rely = 0.49)

		self.preset_manager_button = Button(self.root, text = "Manager", command = self.presetmanager)
		self.preset_manager_button.place(anchor =W, relx = 0.32, rely = 0.49)

		self.preset_manager_button = Button(self.root, text = "Operator", command = self.presetoperator)
		self.preset_manager_button.place(anchor =W, relx = 0.44, rely = 0.49)

		self.submit_button = Button(self.root, text="Create", command=self.create)
		self.submit_button.place(anchor=CENTER, relx=0.35, rely=0.8)

		self.close_button = Button(self.root, text="❌", command=self.close)
		self.close_button.place(anchor=CENTER, relx=0.35, rely=0.9)
		
		

		self.populate_list()

		self.root.bind("<Return>", lambda a:self.create())
		self.root.mainloop()
		return None


	def populate_list(self):
		self.permission_list.delete(0, END)
		for i in range(len(self.options)):
			self.permission_list.insert(END, self.options[i])

			self.permission_list.itemconfig(i, {'fg':self.colour(self.perms[i])})

			


	def toggle_permissions(self):
		for i in range(len(self.options)):
			if self.options[i] == self.permission_list.get(ACTIVE):
				break
		print(i)
		if self.perms[i] == 0:
			self.perms[i] = 1
			self.permission_list.itemconfig(i, {'fg':self.colour(1)})
		else:
			self.perms[i] = 0
			self.permission_list.itemconfig(i, {'fg':self.colour(0)})
			
		
	

	

	def colour(self, val):
		if val == 0:
			return "red"
		else:
			return "green"
	def presetadmin(self):
		for i in range(30):
			self.perms[i] = 1
		self.populate_list()
		return True
	def presetmanager(self):
		self.perms = [0,1,0,1,1,1,1,1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
		self.populate_list()
		return True
	def presetoperator(self):
		self.perms = [0,0,0,0,1,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
		self.populate_list()
		return True	
		

	def close(self):
		self.root.destroy()
		self.root.quit()

	def create(self):
		if self.id_entry.get() == "":
			self.id_entry.focus()
		elif self.password_entry.get() == "":
			self.password_entry.focus()
		elif self.name_entry.get() == "":
			self.name_entry.focus()
		else:
			self.newperms = ""
			for i in range(len(self.perms)):
				self.newperms = self.newperms + str(self.perms[i])
			
								
			self.conn = mysql.connector.connect(
				host=self.host[0],
				user=self.host[1],
				passwd=self.host[2],
				database=self.host[3]
				
			)
			self.c = self.conn.cursor()


			
			self.c.execute("SELECT * FROM user WHERE id = %s", [self.id_entry.get()])
			if len(self.c.fetchall()) > 0:
				self.id_entry.delete(0, END)
				self.id_entry.focus()
				Label(self.root, text = "User with ID already exists.", font = ("Arial", 20), fg = "Red").place(anchor= CENTER, relx =0.39, rely = 0.7)
			else:
				self.c.execute("INSERT INTO user(id, name, password, permissions) VALUES(%s, %s, %s, %s)", [self.id_entry.get(), self.name_entry.get(), self.password_entry.get(), self.newperms])
				self.conn.commit()
				self.close()
							
		return True
class edit_user():
	def __init__(self,b):
		
		self.vonn = sqlite3.connect("LocalData")
		self.v = self.vonn.cursor()
		self.v.execute("SELECT * FROM host_connect_info")
		self.host = self.v.fetchall()[0]
		self.root = Tk()
		self.root.title("Edit User")
		self.root.geometry("750x600+0+0")
	
		self.title_label = Label(self.root, text = "Edit User", font = ("Arial", 35))
		self.title_label.place(anchor=CENTER, relx=0.35, rely=0.15)


		Label(self.root, text = "ID:").place(anchor=E, relx=0.2, rely = 0.3)

		Label(self.root, text = "Password:").place(anchor=E, relx=0.2, rely = 0.37)

		Label(self.root, text = "Name:").place(anchor=E, relx=0.2, rely = 0.44)


		Label(self.root, text = "Permission Presets:").place(anchor=E, relx=0.2, rely = 0.49)





		#entrys

		self.id_entry = Entry(self.root, width = 27)
		self.id_entry.place(anchor=W, relx=0.22, rely = 0.3)
		self.id_entry.focus()

		self.password_entry = Entry(self.root, width = 27)
		self.password_entry.place(anchor=W, relx=0.22, rely = 0.37)

		self.name_entry = Entry(self.root, width = 27)
		self.name_entry.place(anchor=W, relx=0.22, rely = 0.44)
		self.perms = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
		

		self.options = [
		    "Delete Local Data",
		    "Edit Items",
		    "Edit Users",
		    "Edit Accounts",
		    "Transactions",
		    "Local Settings",
		    "Global Settings",
		    "Statisitics"]	
		Label(self.root,text="Permissions").place(anchor=W, relx = 0.65, rely = 0.08)
		self.permission_list = Listbox(self.root, height=30, width = 25)
		self.permission_list.place(anchor = NW, relx = 0.65, rely = 0.1)

		self.search_button = Button(self.root, text = "🔍", command = self.search)
		self.search_button.place(anchor=CENTER, relx=0.6, rely = 0.2953)

		self.toggle_permissions_button = Button(self.root, text="Toggle Selected Permission", command = self.toggle_permissions)
		self.toggle_permissions_button.place(anchor=CENTER, relx= 0.35, rely = 0.6)

		self.preset_admin_button = Button(self.root, text = "Admin", command = self.presetadmin)
		self.preset_admin_button.place(anchor =W, relx = 0.22, rely = 0.49)

		self.preset_manager_button = Button(self.root, text = "Manager", command = self.presetmanager)
		self.preset_manager_button.place(anchor =W, relx = 0.32, rely = 0.49)

		self.preset_manager_button = Button(self.root, text = "Operator", command = self.presetoperator)
		self.preset_manager_button.place(anchor =W, relx = 0.44, rely = 0.49)

		self.submit_button = Button(self.root, text="Save", command=self.save)
		self.submit_button.place(anchor=CENTER, relx=0.35, rely=0.8)

		self.close_button = Button(self.root, text="❌", command=self.close)
		self.close_button.place(anchor=CENTER, relx=0.35, rely=0.9)
		
		self.conn = mysql.connector.connect(
			host=self.host[0],
			user=self.host[1],
			passwd=self.host[2],
			database=self.host[3]
			
		)
		self.c = self.conn.cursor()

		self.populate_list()
		if b != None:
			self.search_direct(b)
		

		self.root.bind("<Return>", lambda a:self.enter())
		self.root.mainloop()

		

		return None


	def populate_list(self):
		self.permission_list.delete(0, END)
		for i in range(len(self.options)):
			self.permission_list.insert(END, self.options[i])

			self.permission_list.itemconfig(i, {'fg':self.colour(self.perms[i])})

			


	def toggle_permissions(self):
		for i in range(len(self.options)):
			if self.options[i] == self.permission_list.get(ACTIVE):
				break
		print(i)
		if self.perms[i] == 0:
			self.perms[i] = 1
			self.permission_list.itemconfig(i, {'fg':self.colour(1)})
		else:
			self.perms[i] = 0
			self.permission_list.itemconfig(i, {'fg':self.colour(0)})
			
		
	

	

	def colour(self, val):
		if val == 0:
			return "red"
		else:
			return "green"
	def presetadmin(self):
		for i in range(30):
			self.perms[i] = 1
		self.populate_list()
		return True
	def presetmanager(self):
		self.perms = [0,1,0,1,1,1,1,1,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
		self.populate_list()
		return True
	def presetoperator(self):
		self.perms = [0,0,0,0,1,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
		self.populate_list()
		return True	
		

	def close(self):
		self.root.destroy()
		self.root.quit()

	def save(self):
		
		self.newperms = ""
		for i in range(len(self.perms)):
			self.newperms = self.newperms + str(self.perms[i])
		
							
		


		
		self.c.execute("SELECT * FROM user WHERE id = %s", [self.id_entry.get()])
		if len(self.c.fetchall()) == 1:
			self.c.execute("UPDATE user SET password = %s, name = %s, permissions = %s WHERE id = %s", [self.password_entry.get(),self.name_entry.get(), self.newperms, self.id_entry.get()])
			self.conn.commit()
			self.close()
						
		else:
			Label(self.root, text = "User with ID does not exist.\nOr Massive Error", font = ("Arial", 20), fg = "Red").place(anchor= CENTER, relx =0.39, rely = 0.7)
			self.id_entry.delete(0, END)
			self.name_entry.delete(0, END)
			self.password_entry.delete(0, END)
			self.id_entry.focus()
			
		return True

	def enter(self):
		if self.id_entry.get() != "" and self.name_entry.get() == "":
			self.search()
		else:
			if self.id_entry.get() == "":
				self.id_entry.focus()
			else:
				if self.password_entry.get() == "":
					self.password_entry.focus()
				else:
					self.save()

	def search(self):
		self.c.execute("SELECT * FROM user WHERE id = %s", [self.id_entry.get()])
		self.data = self.c.fetchall()
		if len(self.data) == 1:
			self.password_entry.insert(0, self.data[0][2])
			self.name_entry.insert(0, self.data[0][1])
			self.perms = [0]*30
			for i in range(30):
				self.perms[i] = int(self.data[0][3][i])

			self.populate_list()
		else:
			self.id_entry.delete(0, END)
			self.name_entry.delete(0, END)
			self.password_entry.delete(0, END)

	def search_direct(self, direct):

		self.c = self.conn.cursor()
		self.c.execute("SELECT * FROM user WHERE id = %s", [direct])
		self.data = self.c.fetchall()
		self.id_entry.insert(0, direct)
		if len(self.data) == 1:
			self.password_entry.insert(0, self.data[0][2])
			self.name_entry.insert(0, self.data[0][1])
			self.perms = [0]*30
			for i in range(30):
				self.perms[i] = int(self.data[0][3][i])
			
			self.populate_list()
		else:
			self.id_entry.delete(0, END)
			self.name_entry.delete(0, END)
			self.password_entry.delete(0, END)
class delete_user():
	def __init__(self, host):
		self.vonn = sqlite3.connect("LocalData")
		self.v = self.vonn.cursor()
		self.v.execute("SELECT * FROM host_connect_info")
		self.host = self.v.fetchall()[0]
		self.conn = mysql.connector.connect(
			host=self.host[0],
			user=self.host[1],
			passwd=self.host[2],
			database=self.host[3]
			
		)
		self.c = self.conn.cursor()
		self.root = Tk()
		self.root.title("Create Item")
		self.root.geometry("500x400+0+0")
	
		self.title_label = Label(self.root, text = "Delete User", font = ("Arial", 35))
		self.title_label.place(anchor=CENTER, relx=0.5, rely=0.15)

		self.id_entry = Entry(self.root, width = 35)
		self.id_entry.place(anchor = CENTER, relx= 0.5, rely = 0.3)
		self.id_entry.focus()

		self.info_label = Label(self.root, text= "ID: \nName: ", justify=RIGHT)
		self.info_label.place(anchor = E, relx=0.5, rely=0.5)

		self.result_label = Label(self.root, text= "", justify=LEFT)
		self.result_label.place(anchor = W, relx=0.5, rely=0.5)
		self.bob = None
		self.root.bind("<Return>", lambda a:self.enter())
		self.root.mainloop()
	def close(self):
		self.root.destroy()
		self.root.quit()

	def enter(self):
		if self.result_label.cget("text") == "":
			self.c.execute("SELECT * FROM user WHERE id = %s", [self.id_entry.get()])
			self.data = self.c.fetchall()
			if len(self.data) == 1:
				self.result_label.configure(text = f'{self.data[0][0]}\n{self.data[0][1]}')
				self.bob = self.data[0][0]
		elif self.bob != self.id_entry.get():
			self.result_label.configure(text = "")
			self.id_entry.delete(0, END

				)

		else:
			if messagebox.askquestion ("Delete User", "Are you sure you want to delete this item?", icon = 'warning') == "yes":
				self.c.execute("DELETE FROM user WHERE id = %s", [self.id_entry.get()])
				self.conn.commit()
				self.close()
			else:
				self.id_entry.delete(0, END)
				self.result_label.configure(text="")

class create_account():
	def __init__(self):
		
		self.vonn = sqlite3.connect("LocalData")
		self.v = self.vonn.cursor()
		self.v.execute("SELECT * FROM host_connect_info")
		self.host = self.v.fetchall()[0]




		self.root = Tk()
		self.root.title("Create Account")
		self.root.geometry("500x400+0+0")
		self.conn = mysql.connector.connect(
			host=self.host[0],
			user=self.host[1],
			passwd=self.host[2],
			database=self.host[3]
									
		)
		self.c = self.conn.cursor()
	
		self.title_label = Label(self.root, text = "Create Account", font = ("Arial", 35))
		self.title_label.place(anchor=CENTER, relx=0.5, rely=0.15)

		self.id_label = Label(self.root, text = "Account ID:")
		self.id_label.place(anchor=E, relx=0.38, rely = 0.3)

		self.name_label = Label(self.root, text = "Name:")
		self.name_label.place(anchor=E, relx=0.38, rely = 0.37)

		self.creditLimit_label = Label(self.root, text = "Credit Limit: $")
		self.creditLimit_label.place(anchor=E, relx=0.38, rely = 0.44)

		self.email_label = Label(self.root, text = "Email:")
		self.email_label.place(anchor=E, relx=0.38, rely = 0.51)

		self.phone_label = Label(self.root, text = "Phone:")
		self.phone_label.place(anchor=E, relx=0.38, rely = 0.57)


		#entrys

		self.id_entry = Entry(self.root)
		self.id_entry.place(anchor=W, relx=0.4, rely = 0.3)
		self.id_entry.focus()

		self.name_entry = Entry(self.root)
		self.name_entry.place(anchor=W, relx=0.4, rely = 0.37)
		
		self.creditLimit_entry = Entry(self.root)
		self.creditLimit_entry.place(anchor=W, relx = 0.4, rely=0.44)

		self.email_entry = Entry(self.root)
		self.email_entry.place(anchor=W, relx=0.4, rely = 0.51)

		self.phone_entry = Entry(self.root)
		self.phone_entry.place(anchor=W, relx=0.4, rely = 0.58)

		self.submit_button = Button(self.root, text="Create", command=self.create)
		self.submit_button.place(anchor=CENTER, relx=0.5, rely=0.8)

		self.close_button = Button(self.root, text="❌", command=self.close)
		self.close_button.place(anchor=CENTER, relx=0.5, rely=0.9)

		

		self.root.bind("<Return>", lambda a:self.create())
		self.root.mainloop()
		return None

	def close(self):
		self.root.destroy()
		self.root.quit()

	def create(self):
		if self.id_entry.get() == "":
			self.id_entry.focus()
		else:
			if self.name_entry.get() == "":
				self.name_entry.focus()
			else:
				if self.creditLimit_entry.get() == "":
					self.creditLimit_entry.focus()
				else:
					if self.email_entry.get() == "":
						self.email_entry.focus()
					else:
						if self.phone_entry.get() == "":
							self.phone_entry.focus()
						else:
							if not basic_operations.is_float(self.creditLimit_entry.get()):
								self.creditLimit_entry.delete(0, END)
								self.creditLimit_entry.focus()
							else:
								
								self.c.execute("SELECT * FROM accounts WHERE id = %s", [self.id_entry.get()])
								self.data1 = self.c.fetchall()
								self.c.execute("SELECT * FROM items WHERE plu = %s", [self.id_entry.get()])
								self.data2 = self.c.fetchall()
								
								if len(self.data1) == 0 and len(self.data2) ==0:
									self.id = self.id_entry.get()
									
									self.c.execute("INSERT INTO accounts(id, name, value, credit_limit, points, email, phone) VALUES(%s, %s, 0, %s,0,%s,%s)",[self.id, self.name_entry.get(), self.creditLimit_entry.get(),self.email_entry.get(), self.phone_entry.get()])
									self.conn.commit()
									
									self.root.destroy()
									self.root.quit()
								else:
									Label(self.root, text="ID number already in use", fg = 'red').place(anchor=CENTER, relx=0.5, rely=0.65)
									self.id_entry.delete(0, END)
									self.id_entry.focus()
				

		return True
class edit_account():
	def __init__(self,ids):
		self.id_direct = ids
		self.vonn = sqlite3.connect("LocalData")
		self.v = self.vonn.cursor()
		self.v.execute("SELECT * FROM host_connect_info")
		self.host = self.v.fetchall()[0]

		
		self.conn = mysql.connector.connect(
									host=self.host[0],
									user=self.host[1],
									passwd=self.host[2],
									database=self.host[3]
								)
		self.c = self.conn.cursor()
		self.root = Tk()
		self.root.title("Edit Item")
		self.root.geometry("500x400+0+0")
	
		self.title_label = Label(self.root, text = "Edit Account", font = ("Arial", 35))
		self.title_label.place(anchor=CENTER, relx=0.5, rely=0.15)

		self.id_label = Label(self.root, text = "Account ID:")
		self.id_label.place(anchor=E, relx=0.38, rely = 0.3)

		self.name_label = Label(self.root, text = "Name:")
		self.name_label.place(anchor=E, relx=0.38, rely = 0.37)

		self.creditLimit_label = Label(self.root, text = "Credit Limit: $")
		self.creditLimit_label.place(anchor=E, relx=0.38, rely = 0.44)

		self.email_label = Label(self.root, text = "Email:")
		self.email_label.place(anchor=E, relx=0.38, rely = 0.51)

		self.phone_label = Label(self.root, text = "Phone:")
		self.phone_label.place(anchor=E, relx=0.38, rely = 0.57)

		self.email_label = Label(self.root, text = "Balance: $")
		self.email_label.place(anchor=E, relx=0.38, rely = 0.65)

		self.phone_label = Label(self.root, text = "Points:")
		self.phone_label.place(anchor=E, relx=0.38, rely = 0.72)


		#entrys

		self.id_entry = Entry(self.root)
		self.id_entry.place(anchor=W, relx=0.4, rely = 0.3)
		self.id_entry.focus()

		self.name_entry = Entry(self.root)
		self.name_entry.place(anchor=W, relx=0.4, rely = 0.37)
		
		self.creditLimit_entry = Entry(self.root)
		self.creditLimit_entry.place(anchor=W, relx = 0.4, rely=0.44)

		self.email_entry = Entry(self.root)
		self.email_entry.place(anchor=W, relx=0.4, rely = 0.51)

		self.phone_entry = Entry(self.root)
		self.phone_entry.place(anchor=W, relx=0.4, rely = 0.58)

		self.value_entry = Entry(self.root)
		self.value_entry.place(anchor=W, relx=0.4, rely = 0.65)

		self.points_entry = Entry(self.root)
		self.points_entry.place(anchor=W, relx=0.4, rely = 0.72)

		self.submit_button = Button(self.root, text="Update", command=self.enter)
		self.submit_button.place(anchor=CENTER, relx=0.5, rely=0.8)

		self.close_button = Button(self.root, text="❌", command=self.close)
		self.close_button.place(anchor=CENTER, relx=0.5, rely=0.9)


		if self.id_direct != None:
			self.id_entry.insert(0, self.id_direct)
			self.search()

		self.root.bind("<Return>", lambda a:self.enter())
		self.root.mainloop()
		return None
	def close(self):
		self.root.destroy()
		self.root.quit()

	def enter(self):
	
		if self.name_entry.get() == "":
			self.name_entry.focus()
		else:
			if self.creditLimit_entry.get() == "":
				self.creditLimit_entry.focus()
			else:
				if self.email_entry.get() == "":
					self.email_entry.focus()
				else:
					if self.phone_entry.get() == "":
						self.phone_entry.focus()
					else:
						
						if basic_operations.is_float(self.creditLimit_entry.get()):
							self.c.execute("UPDATE accounts SET name = %s, value = %s, credit_limit = %s, points =%s, email  = %s, phone = %s WHERE id = %s", [self.name_entry.get(), self.value_entry.get(), self.creditLimit_entry.get(),self.points_entry.get(), self.email_entry.get(), self.phone_entry.get(),self.id_entry.get()])
							self.conn.commit()
							self.root.destroy()
							self.root.quit()
					
			
	def search(self):
		self.c.execute("SELECT * FROM accounts WHERE id = %s", [self.id_entry.get()])
		self.data = self.c.fetchall()
		self.name_entry.insert(0, self.data[0][1])
		self.value_entry.insert(0, self.data[0][2])
		self.creditLimit_entry.insert(0, self.data[0][3])
		self.points_entry.insert(0, self.data[0][4])
		self.email_entry.insert(0, self.data[0][5])
		self.phone_entry.insert(0, self.data[0][6])
	
	def refil_entry(self, entry, inputa):
		entry.delete(0, END)
		entry.insert(0, inputa)


class view_database():
	def __init__(self):
		self.vonn = sqlite3.connect("LocalData")
		self.v = self.vonn.cursor()
		self.v.execute("SELECT * FROM login_info")
		self.data = self.v.fetchall()
		self.perms =self.data[0][2]

		self.v.execute("SELECT * FROM host_connect_info")
		self.host = self.v.fetchall()[0]
		self.conn = mysql.connector.connect(

			host=self.host[0],
			user=self.host[1],
			passwd=self.host[2],
			database=self.host[3]
			
		)
		self.c = self.conn.cursor()
		self.root = Tk()
		self.root.title('Table Viewer')
		self.root.geometry("600x500+0+0")
		self.display = ttk.Notebook(self.root)
		if self.perms[1] == "1":
			self.itemtab = ttk.Frame(self.display)
			self.display.add(self.itemtab, text = 'Items')
		if self.perms[2] == "1":
			self.usertab = ttk.Frame(self.display)
			self.display.add(self.usertab, text = 'Users')
			
		if self.perms[3] == "1":
			self.accounttab = ttk.Frame(self.display)
			self.display.add(self.accounttab, text = 'Accounts')
		if self.perms[4] == "1":
			self.transactiontab = ttk.Frame(self.display)
			self.display.add(self.transactiontab, text = 'Transactions')
		
		
		self.user_checkk = False
		self.display.pack(expand=1, fill='both')

		if self.perms[2] == "1":
			self.setup_user()
			self.append_users(False)
			self.user_table.bind("<Double-Button-1>", lambda a:self.open_user())
			self.user_table.bind("<BackSpace>", lambda a:self.delete_users())
			self.user_table.bind("<Button-1>", lambda a:self.user_check())

		
		self.item_checkk = False
		self.display.pack(expand=1, fill='both')

		if self.perms[1] == "1":
			self.setup_item()
			self.append_items(False)
			self.item_table.bind("<Double-Button-1>", lambda a:self.open_item())
			self.item_table.bind("<BackSpace>", lambda a:self.delete_items())
			self.item_table.bind("<Button-1>", lambda a:self.item_check())

		self.account_checkk = False
		self.display.pack(expand=1, fill='both')

		if self.perms[3] == "1":
			self.setup_account()
			self.append_accounts(False)
			self.account_table.bind("<Double-Button-1>", lambda a:self.open_account())
			self.account_table.bind("<BackSpace>", lambda a:self.delete_accounts())
			self.account_table.bind("<Button-1>", lambda a:self.account_check())

		if self.perms[4] == "1":
			
			self.path_entry = Entry(self.transactiontab, width = 50)
			self.path_entry.place(anchor=CENTER, relx = 0.5, rely=0.35)
			self.path_entry.insert(0, f"{os.path.expanduser('~')}/Desktop")
			

			Button(self.transactiontab, text="Export to excel file", command = self.export_to_excel).place(anchor=CENTER, relx = 0.5, rely=0.5)
			Label(self.transactiontab, text="A folder will be created called JamesPOS which will contain all exported data.").place(anchor=CENTER, relx=0.5, rely=0.7)

		self.root.mainloop()

	#users
	def user_check(self):
		if self.user_checkk:
			self.append_users(True)
			self.user_checkk = False
	def setup_user(self):

		self.user_table_scroll = Scrollbar(self.usertab)
		self.user_table_scroll.pack(side=RIGHT, fill=Y)
			
		
		self.create_button = Button(self.usertab, text = "Create New", fg='blue', command = lambda:self.new_user())
		self.create_button.place(anchor=CENTER, relx=1/6, rely = 0.05)

		self.delete_button = Button(self.usertab, text = "Delete Selected", fg='blue', command=self.delete_users)
		self.delete_button.place(anchor=CENTER, relx=3/6, rely = 0.05)

		self.refresh_button = Button(self.usertab, text = "Refresh", fg='blue', command=lambda:self.append_users(True))
		self.refresh_button.place(anchor=CENTER, relx=5/6, rely = 0.05)


		self.user_table = ttk.Treeview(self.usertab, selectmode="extended", yscrollcommand = self.user_table_scroll.set, height=20)
		self.user_table_scroll.configure(command=self.user_table)
		self.user_table_scroll.configure(command=self.user_table.yview)

		self.headings = [
			('User ID', 
			'User Name', 
			'User Password', 
			'User Permissions'),
			(80,
			80,
			90,
			255)]

		self.user_table['columns'] = (self.headings[0])
		self.user_table.column("#0", anchor=CENTER, width=0, stretch=NO)
		self.user_table.heading("#0", anchor=CENTER)
		for i in range(len(self.headings[0])):
			
			self.user_table.column(self.headings[0][i], anchor=CENTER, minwidth=self.headings[1][i], width=self.headings[1][i], stretch=NO)
			self.user_table.heading(self.headings[0][i], anchor=CENTER, text=self.headings[0][i])
		self.user_table.place(anchor=N, relx=0.48, rely=0.1)
	def new_user(self):
		self.user_checkk=True
		create_user()
	def append_users(self, a):
		if a:
			self.conn = mysql.connector.connect(

			host=self.host[0],
			user=self.host[1],
			passwd=self.host[2],
			database=self.host[3]
			
		)
		
		self.user_table.delete(*self.user_table.get_children())
		
		
		self.c = self.conn.cursor()	
		self.c.execute("SELECT * FROM user")
		self.data = self.c.fetchall()
		
		for i in range(len(self.data)):
			self.user_table.insert(parent='',index='end',iid=i,text='',values=(f'{str(self.data[i][0])}',self.data[i][1],"****",self.data[i][3]))

		return True
	def open_user(self):
		self.curUser = self.user_table.focus()
		self.data = self.user_table.item(self.curUser)['values']
		try:
			self.data[0]
			self.success = True
		except:
			self.success = False
		if self.success:
			while True:
				self.c.execute("SELECT * FROM user WHERE id = %s", [str(self.data[0])])
				self.data2 = self.c.fetchall()
				
				try:
					self.data2[0]
					break
				except IndexError:
					self.data[0] = "0" + str(self.data[0])
			edit_user(self.data[0])
	def delete_users(self):
		self.curUser = self.user_table.focus()
		self.data = self.user_table.item(self.curUser)['values']
		try:
			self.data[0]
			self.success = True
		except:
			self.success = False
		if self.success:
			while True:
				self.c.execute("SELECT * FROM user WHERE id = %s", [str(self.data[0])])
				self.data2 = self.c.fetchall()
				
				try:
					self.data2[0]
					break
				except IndexError:
					self.data[0] = "0" + str(self.data[0])
	
			
			if messagebox.askquestion ("Delete User", f"Are you sure you want to delete this user?\n{self.data[0]}:{self.data[1]}", icon = 'warning') == "yes":
				self.c.execute("DELETE FROM user WHERE id = %s", [self.data[0]])
				self.conn.commit()
				self.append_users(True)
					
				

	#items
	def item_check(self):
		if self.item_checkk:
			self.append_items(True)
			self.item_checkk = False
	def setup_item(self):

		self.item_table_scroll = Scrollbar(self.itemtab)
		self.item_table_scroll.pack(side=RIGHT, fill=Y)
			
		
		self.create_button = Button(self.itemtab, text = "Create New", fg='blue', command = lambda:self.new_item())
		self.create_button.place(anchor=CENTER, relx=1/6, rely = 0.05)

		self.delete_button = Button(self.itemtab, text = "Delete Selected", fg='blue', command=self.delete_items)
		self.delete_button.place(anchor=CENTER, relx=3/6, rely = 0.05)

		self.refresh_button = Button(self.itemtab, text = "Refresh", fg='blue', command=lambda:self.append_items(True))
		self.refresh_button.place(anchor=CENTER, relx=5/6, rely = 0.05)


		self.item_table = ttk.Treeview(self.itemtab, selectmode="extended", yscrollcommand = self.item_table_scroll.set, height=20)
		self.item_table_scroll.configure(command=self.item_table)
		self.item_table_scroll.configure(command=self.item_table.yview)

		self.headings = [
			('PLU', 
			'Description', 
			'Type', 
			'Category',
			'Bonus Points',
			'Price $',
			'Quantity'),
			(72,
			72,
			72,
			72,
			72,
			64,
			80)]

		self.item_table['columns'] = (self.headings[0])
		self.item_table.column("#0", anchor=CENTER, width=0, stretch=NO)
		self.item_table.heading("#0", anchor=CENTER)
		for i in range(len(self.headings[0])):
			
			self.item_table.column(self.headings[0][i], anchor=CENTER, minwidth=self.headings[1][i]-10, width=self.headings[1][i], stretch=NO)
			self.item_table.heading(self.headings[0][i], anchor=CENTER, text=self.headings[0][i])
		self.item_table.place(anchor=N, relx=0.48, rely=0.1)
	def new_item(self):
		self.item_checkk=True
		create_item()
	def append_items(self, a):
		if a:
			self.conn = mysql.connector.connect(

			host=self.host[0],
			user=self.host[1],
			passwd=self.host[2],
			database=self.host[3]
			
		)
		
		self.item_table.delete(*self.item_table.get_children())
		
		
		self.c = self.conn.cursor()	
		self.c.execute("SELECT * FROM items")
		self.data = self.c.fetchall()
		
		for i in range(len(self.data)):
			self.item_table.insert(parent='',index='end',iid=i,text='',values=(f'{str(self.data[i][0])}',self.data[i][1],typeea(self.data[i][2]),self.data[i][3],self.data[i][4],self.data[i][5],self.data[i][6]))

		return True
	def open_item(self):
		
		self.curitem = self.item_table.focus()

		self.data = self.item_table.item(self.curitem)['values']
		try:
			self.data[0]
			self.success = True
		except:
			self.success = False
		if self.success:
			while True:
				self.c.execute("SELECT * FROM items WHERE plu = %s", [str(self.data[0])])
				self.data2 = self.c.fetchall()
				
				try:
					self.data2[0]
					break
				except IndexError:
					self.data[0] = "0" + str(self.data[0])
			
			edit_item(self.data[0])

	def delete_items(self):
		self.curitem = self.item_table.focus()
		self.data = self.item_table.item(self.curitem)['values']
		try:
			self.data[0]
			self.success = True
		except:
			self.success = False
		if self.success:
			while True:
				self.c.execute("SELECT * FROM items WHERE plu = %s", [str(self.data[0])])
				self.data2 = self.c.fetchall()
				try:
					self.data2[0]
					break
				except IndexError:
					self.data[0] = "0" + str(self.data[0])
					
			
			if messagebox.askquestion ("Delete User", f"Are you sure you want to delete this item?\n{self.data[0]}:{self.data[1]}", icon = 'warning') == "yes":
				self.c.execute("DELETE FROM items WHERE plu = %s", [self.data[0]])
				self.conn.commit()
				self.append_items(True)

	#accounts
	def account_check(self):
		if self.account_checkk:
			self.append_accounts(True)
			self.account_checkk = False
	def setup_account(self):

		self.account_table_scroll = Scrollbar(self.accounttab)
		self.account_table_scroll.pack(side=RIGHT, fill=Y)
			
		
		self.create_button = Button(self.accounttab, text = "Create New", fg='blue', command = lambda:self.new_account())
		self.create_button.place(anchor=CENTER, relx=1/6, rely = 0.05)

		self.delete_button = Button(self.accounttab, text = "Delete Selected", fg='blue', command=self.delete_accounts)
		self.delete_button.place(anchor=CENTER, relx=3/6, rely = 0.05)

		self.refresh_button = Button(self.accounttab, text = "Refresh", fg='blue', command=lambda:self.append_accounts(True))
		self.refresh_button.place(anchor=CENTER, relx=5/6, rely = 0.05)


		self.account_table = ttk.Treeview(self.accounttab, selectmode="extended", yscrollcommand = self.account_table_scroll.set, height=20)
		self.account_table_scroll.configure(command=self.account_table)
		self.account_table_scroll.configure(command=self.account_table.yview)

		self.headings = [
			('ID', 
			'Name', 
			'Value', 
			'Credit Limit',
			'Points',
			'Email',
			'Phone'),
			(72,
			72,
			72,
			72,
			72,
			64,
			80)]

		self.account_table['columns'] = (self.headings[0])
		self.account_table.column("#0", anchor=CENTER, width=0, stretch=NO)
		self.account_table.heading("#0", anchor=CENTER)
		for i in range(len(self.headings[0])):
			
			self.account_table.column(self.headings[0][i], anchor=CENTER, minwidth=self.headings[1][i], width=self.headings[1][i], stretch=NO)
			self.account_table.heading(self.headings[0][i], anchor=CENTER, text=self.headings[0][i])
		self.account_table.place(anchor=N, relx=0.48, rely=0.1)
	def new_account(self):
		self.account_checkk=True
		create_account()
	def append_accounts(self, a):
		if a:
			self.conn = mysql.connector.connect(

			host=self.host[0],
			user=self.host[1],
			passwd=self.host[2],
			database=self.host[3]
			
		)
		
		self.account_table.delete(*self.account_table.get_children())
		
		
		self.c = self.conn.cursor()	
		self.c.execute("SELECT * FROM accounts")
		self.data = self.c.fetchall()
		
		for i in range(len(self.data)):
			self.account_table.insert(parent='',index='end',iid=i,text='',values=(f'{str(self.data[i][0])}',self.data[i][1],self.data[i][2],self.data[i][3],self.data[i][4],self.data[i][5],self.data[i][6]))

		return True
	def open_account(self):
		
		self.curaccount = self.account_table.focus()
		self.data = self.account_table.item(self.curaccount)['values']
		try:
			self.data[0]
			self.success = True
		except:
			self.success = False
		if self.success:
			while True:
				self.c.execute("SELECT * FROM accounts WHERE id = %s", [str(self.data[0])])
				self.data2 = self.c.fetchall()
				
				try:
					self.data2[0]
					break
				except IndexError:
					self.data[0] = "0" + str(self.data[0])
			edit_account(self.data[0])
		
	
	def delete_accounts(self):
		self.curaccount = self.account_table.focus()
		self.data = self.account_table.item(self.curaccount)['values']
		try:
			self.data[0]
			self.success = True
		except:
			self.success = False
		if self.success:
			while True:
				self.c.execute("SELECT * FROM accounts WHERE id = %s", [str(self.data[0])])
				self.data2 = self.c.fetchall()
				try:
					self.data2[0]
					break
				except IndexError:
					self.data[0] = "0" + str(self.data[0])

		if messagebox.askquestion ("Delete User", f"Are you sure you want to delete this account?\n{self.data[0]}:{self.data[1]}", icon = 'warning') == "yes":
			self.c.execute("DELETE FROM accounts WHERE id = %s", [self.data[0]])
			self.conn.commit()
			self.append_accounts(True)
	def export_to_excel(self):
		self.datetime = datetime.datetime.now()
		try:
			self.path = f"{self.path_entry.get()}/JamesPOS"
			os.mkdir(self.path)
			
		except FileExistsError:
			pass
		try:
			self.path = f"{self.path_entry.get()}/JamesPOS/transactions"
			os.mkdir(self.path)
		except FileExistsError:
			pass
	
		self.wb = Workbook()
		self.tranlist = self.wb.create_sheet("Transaction List", 0)
		
		self.tranlist.append(['ID', 'Date', 'Time', 'Total(gst exc.)', 'GST', 'Payment Method', 'Loyalty ID', 'Discounts', 'Amount Recieved', 'Location', 'User ID'])
		self.tranlist["A1"].font = Font(bold=True)
		self.tranlist["B1"].font = Font(bold=True)
		self.tranlist["C1"].font = Font(bold=True)
		self.tranlist["D1"].font = Font(bold=True)
		self.tranlist["E1"].font = Font(bold=True)
		self.tranlist["F1"].font = Font(bold=True)
		self.tranlist["G1"].font = Font(bold=True)
		self.tranlist["H1"].font = Font(bold=True)
		self.tranlist["I1"].font = Font(bold=True)
		self.tranlist["J1"].font = Font(bold=True)
		self.tranlist["K1"].font = Font(bold=True)

		self.c.execute("SELECT * FROM transactions ORDER BY id DESC")
		self.data = self.c.fetchall()

		for i in range(len(self.data)):
			self.tranlist.append([self.data[i][0], self.data[i][1], self.data[i][2], self.data[i][3], self.data[i][4], self.data[i][5], self.data[i][6], self.data[i][7], self.data[i][8], self.data[i][9], self.data[i][10]])

		self.tranitems = self.wb.create_sheet("Transaction Items", 1)
		self.tranitems.append(['ID', 'PLU', 'Description', 'Price(gst exc.)', 'GST', 'Shelf Price', 'Loyalty points'])
		self.tranitems["A1"].font = Font(bold=True)
		self.tranitems["B1"].font = Font(bold=True)
		self.tranitems["C1"].font = Font(bold=True)
		self.tranitems["D1"].font = Font(bold=True)
		self.tranitems["E1"].font = Font(bold=True)
		self.tranitems["F1"].font = Font(bold=True)
		self.tranitems["G1"].font = Font(bold=True)
		

		self.c.execute("SELECT * FROM transaction_items ORDER BY id DESC")
		self.data = self.c.fetchall()

		for i in range(len(self.data)):
			self.tranitems.append([self.data[i][0], self.data[i][1], self.data[i][2], self.data[i][3], self.data[i][4], self.data[i][5],self.data[i][6]])


		self.date = str(self.datetime.strftime('%x'))
		self.date= f'{self.date[0]}{self.date[1]}.{self.date[3]}{self.date[4]}.{self.date[6]}{self.date[7]}'

		self.wb.save(f"{self.path}/Exported {self.datetime.strftime('%c')}.xlsx")



		
		return None
#view_database()






